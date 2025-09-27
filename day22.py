import time
from selenium import webdriver
from selenium.webdriver.common.by import By

from openpyxl import load_workbook

workbook = load_workbook('testdata.xlsx')
sheet = workbook.active
driver = webdriver.Firefox()
driver.maximize_window()
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    us = row[0]
    pw = row[1]
    driver.get("https://www.saucedemo.com/v1/")
    time.sleep(5)
    driver.find_element(By.ID,value='user-name').send_keys(us)
    driver.find_element(By.ID,value='password').send_keys(pw)
    driver.find_element(By.ID,value='login-button').click()
    time.sleep(5)
    driver.find_element(By.XPATH,value="//button[normalize-space()='Open Menu']").click()
    driver.find_element(By.XPATH,value="//a[@id='logout_sidebar_link']").click()
    time.sleep(5)
driver.quit()